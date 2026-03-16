"""Export router — Excel, PDF, CSV exports behind paywall."""

import io
from datetime import datetime

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse

from backend.auth import require_auth, require_plan
from backend.database import get_db

router = APIRouter(prefix="/api/export", tags=["exports"])


@router.get("/company/{cik}/xlsx")
def export_company_xlsx(cik: str, request: Request):
    require_plan(request, "pro")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    with get_db() as conn:
        company = conn.execute("SELECT * FROM companies WHERE cik = ?", (cik,)).fetchone()
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")

        subs = conn.execute(
            "SELECT * FROM subsidiaries WHERE cik = ? ORDER BY sub_name", (cik,)
        ).fetchall()

    wb = Workbook()

    # Company info sheet
    ws_info = wb.active
    ws_info.title = "Company Info"
    header_font = Font(bold=True, size=12)
    ws_info.append(["SubTrack Export"])
    ws_info["A1"].font = Font(bold=True, size=16)
    ws_info.append([])
    ws_info.append(["CIK", company["cik"]])
    ws_info.append(["Company Name", company["company_name"]])
    ws_info.append(["Total Subsidiaries", company["num_subsidiaries"]])
    ws_info.append(["Filing Period", "{} to {}".format(company["first_filing"], company["last_filing"])])
    ws_info.append(["Total Filings", company["num_filings"]])
    ws_info.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])

    # Subsidiaries sheet
    ws_subs = wb.create_sheet("Subsidiaries")
    headers = ["Name", "First Seen", "Last Seen", "Time In", "Time Out", "Confidence", "Type", "Source", "Enriched"]
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws_subs.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, sub in enumerate(subs, 2):
        ws_subs.cell(row=i, column=1, value=sub["sub_name"])
        ws_subs.cell(row=i, column=2, value=sub["first_seen"])
        ws_subs.cell(row=i, column=3, value=sub["last_seen"])
        ws_subs.cell(row=i, column=4, value=sub["time_in"])
        ws_subs.cell(row=i, column=5, value=sub["time_out"])
        ws_subs.cell(row=i, column=6, value=sub["confidence"])
        ws_subs.cell(row=i, column=7, value=sub["type"] or "")
        ws_subs.cell(row=i, column=8, value=sub["source"])
        ws_subs.cell(row=i, column=9, value="Yes" if sub["enriched"] else "No")

    # Auto-width columns
    for ws in [ws_info, ws_subs]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    # Freeze header row
    ws_subs.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    safe_name = quote(company["company_name"].replace(" ", "_")[:30])
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}.xlsx"},
    )


@router.get("/company/{cik}/pdf")
def export_company_pdf(cik: str, request: Request):
    require_plan(request, "pro")
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    with get_db() as conn:
        company = conn.execute("SELECT * FROM companies WHERE cik = ?", (cik,)).fetchone()
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")

        subs = conn.execute(
            "SELECT * FROM subsidiaries WHERE cik = ? ORDER BY sub_name LIMIT 200", (cik,)
        ).fetchall()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("SubTrack Report: {}".format(company["company_name"]), styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("CIK: {} | Subsidiaries: {} | Period: {} to {}".format(
        company["cik"], company["num_subsidiaries"], company["first_filing"], company["last_filing"]
    ), styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Table
    data = [["Subsidiary", "First Seen", "Last Seen", "Time In", "Time Out", "Confidence"]]
    for sub in subs:
        data.append([
            sub["sub_name"][:40],
            sub["first_seen"],
            sub["last_seen"],
            (sub["time_in"] or "")[:35],
            (sub["time_out"] or "")[:35],
            sub["confidence"],
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4ff")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Generated by SubTrack on {}".format(datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        styles["Normal"],
    ))

    doc.build(elements)
    buf.seek(0)

    from urllib.parse import quote
    safe_name = quote(company["company_name"].replace(" ", "_")[:30])
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}.pdf"},
    )


@router.get("/company/{cik}/csv")
def export_company_csv(cik: str, request: Request):
    require_plan(request, "pro")

    with get_db() as conn:
        company = conn.execute("SELECT * FROM companies WHERE cik = ?", (cik,)).fetchone()
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")

        subs = conn.execute(
            "SELECT * FROM subsidiaries WHERE cik = ? ORDER BY sub_name", (cik,)
        ).fetchall()

    def gen():
        yield "CIK,Company,Subsidiary,First Seen,Last Seen,Time In,Time Out,Confidence,Type,Source\n"
        for s in subs:
            name = s["sub_name"].replace('"', '""')
            ti = (s["time_in"] or "").replace('"', '""')
            to = (s["time_out"] or "").replace('"', '""')
            yield '"{}","{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(
                cik, company["company_name"], name,
                s["first_seen"], s["last_seen"], ti, to,
                s["confidence"], s["type"] or "", s["source"]
            )

    from urllib.parse import quote
    safe_name = quote(company["company_name"].replace(" ", "_")[:30])
    return StreamingResponse(
        gen(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}.csv"},
    )


@router.get("/bulk/xlsx")
def export_bulk_xlsx(request: Request):
    require_plan(request, "enterprise")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "All Subsidiaries"

    headers = ["CIK", "Company", "Subsidiary", "First Seen", "Last Seen",
               "Time In", "Time Out", "Confidence", "Type", "Source", "Enriched"]
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    with get_db() as conn:
        row_num = 2
        offset = 0
        batch = 5000
        while True:
            rows = conn.execute(
                """SELECT s.*, c.company_name FROM subsidiaries s
                   JOIN companies c ON s.cik = c.cik
                   ORDER BY c.company_name, s.sub_name
                   LIMIT ? OFFSET ?""",
                (batch, offset),
            ).fetchall()
            if not rows:
                break
            for r in rows:
                ws.cell(row=row_num, column=1, value=r["cik"])
                ws.cell(row=row_num, column=2, value=r["company_name"])
                ws.cell(row=row_num, column=3, value=r["sub_name"])
                ws.cell(row=row_num, column=4, value=r["first_seen"])
                ws.cell(row=row_num, column=5, value=r["last_seen"])
                ws.cell(row=row_num, column=6, value=r["time_in"])
                ws.cell(row=row_num, column=7, value=r["time_out"])
                ws.cell(row=row_num, column=8, value=r["confidence"])
                ws.cell(row=row_num, column=9, value=r["type"] or "")
                ws.cell(row=row_num, column=10, value=r["source"])
                ws.cell(row=row_num, column=11, value="Yes" if r["enriched"] else "No")
                row_num += 1
            offset += batch

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=subtrack_full_export.xlsx"},
    )
